import openpyxl as xl
from openpyxl.chart import BarChart, Reference

# Load the existing Excel file (Make sure 'transactions.xlsx' is in the same folder)
wb = xl.load_workbook('../codes/transaction3.xlsx')
# Select the specific worksheet to work on
sheet = wb['Sheet1']

# Loop through every row starting from row 2 (skipping the header) to the end of the sheet
for row in range(2, sheet.max_row + 1):
    # Grab the cell in the 3rd column (e.g., Original Price)
    cell = sheet.cell(row, 3)

    # Calculate the new price with a 10% discount (multiplying by 0.9)
    corrected_price = cell.value * 0.9

    # Select the cell in the 4th column of the same row to store the result
    corrected_price_cell = sheet.cell(row, 4)

    # Write the new discounted price into that cell
    corrected_price_cell.value = corrected_price

# Define the range of data we want to include in our chart (Column 4: Corrected Prices)
values = Reference(sheet,
                   min_row=2,
                   max_row=sheet.max_row,
                   min_col=4,
                   max_col=4)

# Initialize a new BarChart object
chart = BarChart()
# Attach the data we defined above to the chart
chart.add_data(values)
# Place the chart on the spreadsheet starting at cell E2
sheet.add_chart(chart, "e2")

# Save the modifications into a NEW file so we don't overwrite the original data
wb.save('transaction5.xlsl')